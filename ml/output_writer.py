# ml/output_writer.py
# Generates JSON and Excel outputs from NER results

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


TYPE_COLORS = {
    'MON': 'FFCDD2',  # red-ish  (money)
    'PCT': 'FFF9C4',  # yellow   (percent)
    'DAT': 'BBDEFB',  # blue     (date)
    'TIM': 'B2EBF2',  # cyan     (time)
    'CNT': 'C8E6C9',  # green    (count)
    'FRC': 'E1BEE7',  # purple   (fraction)
    'ORD': 'D7CCC8',  # brown    (ordinal)
    'APX': 'F5F5F5',  # gray     (approximate)
}


def write_json(results: list, meta: dict, output_path: str) -> str:
    type_counts = {}
    for r in results:
        t = r['type']
        type_counts[t] = type_counts.get(t, 0) + 1

    payload = {
        'meta':     meta,
        'count':    len(results),
        'by_type':  type_counts,
        'items':    [
            {
                'id':        r['id'],
                'type':      r['type'],
                'raw':       r['raw'],
                'formatted': r['formatted'],
                'value':     str(r.get('value', '')),
                'unit':      r.get('unit', ''),
                'sent_idx':  r.get('sent_idx', 0),
            }
            for r in results
        ],
    }

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return output_path


def write_excel(results: list, meta: dict, output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Son iboralar'

    # ── Meta sheet ──────────────────────────────────────────
    ws_meta = wb.create_sheet('Meta')
    ws_meta.append(['Key', 'Value'])
    for k, v in meta.items():
        ws_meta.append([k, str(v)])
    for col in ws_meta.columns:
        ws_meta.column_dimensions[get_column_letter(col[0].column)].width = 24

    # ── Main sheet header ───────────────────────────────────
    headers = ['#', 'Type', 'Raw expression', 'Normalized value', 'Unit', 'Sentence #']
    header_fill = PatternFill(fill_type='solid', fgColor='37474F')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin         = Side(style='thin', color='CCCCCC')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell             = ws.cell(row=1, column=col_idx, value=header)
        cell.fill        = header_fill
        cell.font        = header_font
        cell.alignment   = Alignment(horizontal='center', vertical='center')
        cell.border      = border

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # ── Data rows ────────────────────────────────────────────
    for r in results:
        row_n  = r['id'] + 1
        color  = TYPE_COLORS.get(r['type'], 'FFFFFF')
        fill   = PatternFill(fill_type='solid', fgColor=color)

        values = [
            r['id'],
            r['type'],
            r['raw'],
            r['formatted'],
            r.get('unit', ''),
            r.get('sent_idx', 0) + 1,
        ]
        for col_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=row_n, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 2:
                cell.font = Font(bold=True)

    # ── Column widths ────────────────────────────────────────
    col_widths = [6, 8, 36, 28, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Stats sheet ──────────────────────────────────────────
    ws_stats = wb.create_sheet('Statistics')
    ws_stats.append(['Entity type', 'Count', 'Color'])
    type_counts = {}
    for r in results:
        t = r['type']
        type_counts[t] = type_counts.get(t, 0) + 1

    for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        row = ws_stats.append([t, cnt, ''])
        color  = TYPE_COLORS.get(t, 'FFFFFF')
        fill   = PatternFill(fill_type='solid', fgColor=color)
        last_r = ws_stats.max_row
        ws_stats.cell(last_r, 3).fill = fill

    ws_stats.column_dimensions['A'].width = 16
    ws_stats.column_dimensions['B'].width = 10
    ws_stats.column_dimensions['C'].width = 12

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
